"""Client helpers for microservice interactions."""

from __future__ import annotations

import base64
import json
import os
import re
import urllib.parse
from pathlib import Path
from typing import Any

import httpx
from dotenv import load_dotenv

# Ensure local `.env` variables are available even if config hasn't been imported yet.
load_dotenv(Path(__file__).resolve().parent.parent.parent / ".env")

_DEFAULT_CONNECT_TIMEOUT = 60.0
_DEFAULT_READ_TIMEOUT = 300.0


def _parse_timeout(raw: str | None, fallback: float) -> float:
    if not raw:
        return fallback
    try:
        value = float(raw)
    except (TypeError, ValueError):
        return fallback
    if value <= 0:
        return fallback
    return value


def _build_timeout() -> httpx.Timeout:
    connect = _parse_timeout(os.getenv("MICROSERVICE_CONNECT_TIMEOUT"), _DEFAULT_CONNECT_TIMEOUT)
    read = _parse_timeout(os.getenv("MICROSERVICE_READ_TIMEOUT"), _DEFAULT_READ_TIMEOUT)
    # Apply the same value to write timeout to avoid half-open uploads on slow links.
    return httpx.Timeout(timeout=None, connect=connect, read=read, write=read)


def _is_crm_listings_excel(path: Path) -> bool:
    if path.suffix.lower() not in {".xls", ".xlsx"}:
        return False
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return False

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False

    try:
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return False
        normalized = {str(cell).strip().lower() for cell in header_row if cell is not None and str(cell).strip()}
        return {"здание", "тип использования", "площадь, кв.м."}.issubset(normalized)
    finally:
        wb.close()


async def process_file(file_path: Path, chat_id: str) -> tuple[bytes | None, str | None, list[dict[str, Any]], str | None]:
    """
    Send file to microservice.

    Returns a tuple ``(binary_bytes, filename, status_messages, text_message)`` where:
      * ``binary_bytes``/``filename`` contain the attachment returned by MS (e.g. Excel);
      * ``text_message`` contains a textual representation (e.g. CRM response) when MS returns JSON;
      * ``status_messages`` propagates informational messages encoded in `X-UD-Status`.
    Only one of ``binary_bytes``/``filename`` or ``text_message`` will be non-empty.
    """

    base_url = os.getenv("MICROSERVICE_BASE_URL", "").rstrip("/")
    if not base_url:
        raise RuntimeError("MICROSERVICE_BASE_URL is not set")

    url = f"{base_url}/process_file"

    timeout = _build_timeout()
    async with httpx.AsyncClient(timeout=timeout, trust_env=False) as client:
        with file_path.open("rb") as f:
            files = {"file": (file_path.name, f, "application/octet-stream")}
            data = {"chat_id": chat_id}
            if _is_crm_listings_excel(file_path):
                data["crm_forward"] = "1"
            resp = await client.post(url, data=data, files=files)
            resp.raise_for_status()

            cd = resp.headers.get("Content-Disposition") or resp.headers.get("content-disposition") or ""
            filename = _filename_from_content_disposition(cd) or "result.xlsx"
            content_type = (resp.headers.get("Content-Type") or resp.headers.get("content-type") or "").split(";", 1)[0].strip().lower()

            status_messages: list[dict[str, Any]] = []
            header_val = resp.headers.get("X-UD-Status") or resp.headers.get("x-ud-status")
            if header_val:
                try:
                    decoded = base64.b64decode(header_val)
                    data_obj = json.loads(decoded.decode("utf-8"))
                    if isinstance(data_obj, list):
                        status_messages = [entry for entry in data_obj if isinstance(entry, dict)]
                    elif isinstance(data_obj, dict):
                        inner = data_obj.get("status_messages")
                        if isinstance(inner, list):
                            status_messages = [entry for entry in inner if isinstance(entry, dict)]
                        else:
                            status_messages = [data_obj]
                    else:
                        status_messages = [{"message": str(data_obj)}]
                except Exception:
                    status_messages = [{"message": str(header_val)}]

            is_json_response = content_type == "application/json" or not cd
            if is_json_response:
                text_message = _format_json_response(resp)
                return None, None, status_messages, text_message

            content = resp.content
            return content, filename, status_messages, None


_FILENAME_STAR_RE = re.compile(r"filename\*=(?:UTF-8'')?([^;]+)", flags=re.IGNORECASE)
_FILENAME_RE = re.compile(r"filename=\"?([^\";]+)\"?", flags=re.IGNORECASE)


def _filename_from_content_disposition(value: str) -> str | None:
    if not value:
        return None
    m = _FILENAME_STAR_RE.search(value)
    if m:
        raw = m.group(1)
        try:
            return urllib.parse.unquote(raw)
        except Exception:
            return raw
    m = _FILENAME_RE.search(value)
    if m:
        return m.group(1)
    return None


def _format_json_response(resp: httpx.Response) -> str:
    try:
        data = resp.json()
    except ValueError:
        return resp.text

    def _val(obj: Any, default: str = 'нет данных') -> str:
        return str(obj) if obj is not None else default

    crm = data.get('crm_response') if isinstance(data, dict) else None
    if isinstance(crm, dict):
        rid = crm.get('request_id') or data.get('request_id')
        summary = crm.get('summary') if isinstance(crm.get('summary'), dict) else {}
        lines: list[str] = []
        if rid:
            lines.append(f"CRM запрос: {rid}")
        lines.append(
            'Обновлено: {updated}; добавлено: {inserted}; пропущено: {skipped}'.format(
                updated=_val(summary.get('updated'), '0'),
                inserted=_val(summary.get('inserted'), '0'),
                skipped=_val(summary.get('skipped'), '0'),
            )
        )
        duplicates = crm.get('duplicates') if isinstance(crm.get('duplicates'), list) else []
        if duplicates:
            lines.append('Дубликаты:')
            for dup in duplicates[:10]:
                if not isinstance(dup, dict):
                    continue
                idx = dup.get('listing_index')
                reason = dup.get('reason')
                if idx is None and not reason:
                    continue
                if idx is None:
                    lines.append(f"- {reason}")
                elif reason:
                    lines.append(f"- #{idx}: {reason}")
                else:
                    lines.append(f"- #{idx}")
            if len(duplicates) > 10:
                lines.append(f"- ... ещё {len(duplicates) - 10}")
        meta = data.get('meta')
        if isinstance(meta, dict):
            listings_total = meta.get('listings_total')
            if listings_total is not None:
                lines.append(f"Всего строк в запросе: {listings_total}")
        sheet_url = crm.get('sheet_url') or data.get('sheet_url')
        if isinstance(sheet_url, str) and sheet_url.strip():
            lines.append(f"Google Sheet: {sheet_url.strip()}")
        return '\n'.join(lines)

    if isinstance(data, dict):
        return json.dumps(data, ensure_ascii=False, indent=2)
    if isinstance(data, list):
        return json.dumps(data, ensure_ascii=False, indent=2)
    return str(data)

async def get_health() -> dict[str, Any]:
    """Fetch health status from the microservice."""

    base_url = os.getenv("MICROSERVICE_BASE_URL", "").rstrip("/")
    if not base_url:
        raise RuntimeError("MICROSERVICE_BASE_URL is not set")

    url = f"{base_url}/healthz"
    timeout = _build_timeout()
    async with httpx.AsyncClient(timeout=timeout, trust_env=False) as client:
        resp = await client.get(url)
        resp.raise_for_status()
        try:
            return resp.json()
        except json.JSONDecodeError:
            return {"ok": False, "raw": resp.text}
