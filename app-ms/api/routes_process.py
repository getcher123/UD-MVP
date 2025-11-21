from __future__ import annotations

import base64
import json
import os
import time
from pathlib import Path
from typing import Any, Mapping, Optional

from fastapi import APIRouter, UploadFile, File, Form, Response
from fastapi.responses import JSONResponse

from core.config import get_settings
from core.config_loader import get_rules
from core.errors import ErrorCode, ServiceError
from core.ids import new_job_id
from services.agentql_client import run_agentql
from services.audio_client import transcribe_audio
from services.chatgpt_structured import extract_structured_objects
from services.chatgpt_vision import analyze_page_image
from services.crm_client import send_listings_to_crm
from services.crm_payload import prepare_crm_payload
from services.excel_export import build_xlsx
from services.excel_to_csv import excel_to_csv_text
from services.docx_to_md import docx_to_md_text
from services.ppt_to_md import ppt_to_md_text
from services.listings import flatten_objects_to_listings
from services.normalize import normalize_agentql_payload
from services.pdf_convert import to_pdf
from services.pdf_to_images import pdf_to_images
from utils.fs import (
    build_result_path,
    enforce_size_limit,
    file_ext,
    is_allowed_type,
    write_bytes,
    write_text,
)


router = APIRouter(tags=["process"])
settings = get_settings()


def _persist_json(data: dict, req_id: str, name: str) -> None:
    try:
        target = build_result_path(req_id, name, base_dir=settings.RESULTS_DIR)
        write_text(target, json.dumps(data, ensure_ascii=False, indent=2))
    except Exception:  # pragma: no cover - diagnostics helper
        pass


def _get_pipeline_cfg(rules: Mapping[str, Any]) -> Mapping[str, Any]:
    pipeline = rules.get("pipeline") if isinstance(rules, Mapping) else None
    return pipeline if isinstance(pipeline, Mapping) else {}


def _get_format_cfg(pipeline_cfg: Mapping[str, Any], fmt: str) -> Mapping[str, Any]:
    raw = pipeline_cfg.get(fmt) if isinstance(pipeline_cfg, Mapping) else None
    return raw if isinstance(raw, Mapping) else {}


def _get_stage_cfg(pipeline_cfg: Mapping[str, Any], fmt: str, stage: str) -> Mapping[str, Any]:
    cfg: dict[str, Any] = {}
    common = pipeline_cfg.get("common") if isinstance(pipeline_cfg, Mapping) else None
    if isinstance(common, Mapping):
        stage_common = common.get(stage)
        if isinstance(stage_common, Mapping):
            cfg.update(stage_common)
    fmt_cfg = pipeline_cfg.get(fmt) if isinstance(pipeline_cfg, Mapping) else None
    if isinstance(fmt_cfg, Mapping):
        stage_specific = fmt_cfg.get(stage)
        if isinstance(stage_specific, Mapping):
            cfg.update(stage_specific)
    return cfg


def _cfg_enabled(cfg: Mapping[str, Any] | None, default: bool = True) -> bool:
    if not isinstance(cfg, Mapping):
        return default
    enabled = cfg.get("enabled")
    if enabled is None:
        return default
    if isinstance(enabled, bool):
        return enabled
    if isinstance(enabled, (int, float)):
        return bool(enabled)
    if isinstance(enabled, str):
        return enabled.strip().lower() not in {"0", "false", "no", "off"}
    return default


def _detect_format(ext: str, is_audio: bool) -> str:
    if is_audio:
        return "audio"
    if ext == "docx":
        return "docx"
    if ext == "doc":
        return "doc"
    if ext in {"ppt", "pptx"}:
        return "ppt"
    if ext in {"xls", "xlsx", "xlsm"}:
        return "excel"
    if ext == "pdf":
        return "pdf"
    if ext in {"jpg", "jpeg", "png"}:
        return "image"
    if ext == "txt":
        return "txt"
    return "doc"


@router.post("/process_file")
async def process_file(
    file: UploadFile = File(...),
    query: Optional[str] = Form(None),
    request_id: Optional[str] = Form(None),
    output: Optional[str] = Form("excel"),
    crm_forward: Optional[str] = Form(None),
) -> dict:
    start_ts = time.perf_counter()
    req_id = request_id or new_job_id()
    raw_filename = file.filename or "upload"
    filename_candidate = Path(raw_filename).name
    filename = filename_candidate if filename_candidate else "upload"

    src_path = build_result_path(req_id, filename, base_dir=settings.RESULTS_DIR)
    data = await file.read()
    write_bytes(src_path, data)

    try:
        enforce_size_limit(src_path, settings.MAX_FILE_MB)
    except ValueError as e:
        raise ServiceError(ErrorCode.VALIDATION_ERROR, 400, str(e))
    if not is_allowed_type(src_path, settings.ALLOW_TYPES):
        raise ServiceError(ErrorCode.UNSUPPORTED_TYPE, 400, f"Unsupported file type: {Path(src_path).suffix}")

    ext = file_ext(src_path)
    is_audio = ext in settings.AUDIO_TYPES
    is_doc = ext == "doc"
    is_docx = ext in settings.DOCX_TYPES
    is_excel = ext in settings.EXCEL_TYPES
    is_ppt = ext in {"ppt", "pptx"}

    rules = get_rules(settings.RULES_PATH)
    pipeline_cfg = _get_pipeline_cfg(rules)
    fmt_key = _detect_format(ext, is_audio)
    treat_as_pdf = fmt_key in {"pdf", "image"}
    format_cfg = _get_format_cfg(pipeline_cfg, fmt_key)

    stem_lower = Path(filename).stem.lower()
    force_crm = crm_forward == "1"
    crm_candidate = force_crm
    crm_payload: dict[str, Any] | None = None
    if not crm_candidate and ext in {"xls", "xlsx"}:
        if stem_lower.startswith("listing") or _excel_has_listing_headers(src_path, rules):
            crm_candidate = True

    if crm_candidate:
        try:
            crm_payload = prepare_crm_payload(str(src_path), req_id, filename, rules)
        except ServiceError as exc:
            if force_crm or stem_lower.startswith("listing"):
                raise
            if exc.code != ErrorCode.CRM_SYNC_ERROR:
                raise
            crm_payload = None
            crm_candidate = False

    if crm_payload:
        _persist_json(crm_payload, req_id, "crm_request.json")
        crm_response = send_listings_to_crm(crm_payload, settings)
        _persist_json(crm_response, req_id, "crm_response.json")
        elapsed_ms = int((time.perf_counter() - start_ts) * 1000)

        sheet_url = crm_response.get("sheet_url")
        body = {
            "request_id": req_id,
            "crm_response": crm_response,
            "meta": {
                "source_file": os.path.basename(filename),
                "listings_total": len(crm_payload.get("listings", [])),
                "timing_ms": elapsed_ms,
                "pipeline": "crm_forward",
            },
        }
        if isinstance(sheet_url, str) and sheet_url.strip():
            body["crm_response"]["sheet_url"] = sheet_url.strip()
        return JSONResponse(content=body)

    pipeline_steps: list[str] = []
    if fmt_key:
        pipeline_steps.append(fmt_key)

    payload: dict[str, object]
    pdf_path: str | None = None
    pending_questions: list[dict[str, object]] = []
    agentql_mode_meta: Optional[str] = None

    if is_audio:
        transcription_cfg = _get_stage_cfg(pipeline_cfg, "audio", "transcription")
        if not _cfg_enabled(transcription_cfg, True):
            raise ServiceError(ErrorCode.INTERNAL_ERROR, 503, "Audio transcription disabled via configuration")

        transcription = transcribe_audio(data, filename, settings)
        _persist_json(transcription, req_id, "app_audio_response.json")

        srt = transcription.get("srt") if isinstance(transcription, dict) else None
        if not isinstance(srt, str) or not srt.strip():
            raise ServiceError(ErrorCode.TRANSCRIPTION_ERROR, 424, "app-audio returned empty SRT payload")

        srt_name = f"{Path(filename).stem or 'audio'}.srt"
        try:
            write_text(build_result_path(req_id, srt_name, base_dir=settings.RESULTS_DIR), srt)
        except Exception:  # pragma: no cover - diagnostics helper
            pass

        pipeline_steps.append("transcription")

        chatgpt_cfg = _get_stage_cfg(pipeline_cfg, "audio", "chatgpt_structured")
        if _cfg_enabled(chatgpt_cfg, True):
            payload = extract_structured_objects(srt)
            pipeline_steps.append("chatgpt_structured")
            _persist_json(payload, req_id, "chatgpt_structured.json")
        else:
            payload = {"objects": []}
            pipeline_steps.append("chatgpt_skip")
    elif is_excel:
        convert_cfg = _get_stage_cfg(pipeline_cfg, "excel", "excel_to_csv")
        if not _cfg_enabled(convert_cfg, True):
            raise ServiceError(ErrorCode.INTERNAL_ERROR, 503, "Excel to CSV disabled via configuration")

        try:
            csv_text = excel_to_csv_text(src_path)
        except Exception as exc:  # noqa: BLE001
            raise ServiceError(ErrorCode.INTERNAL_ERROR, 500, f"Failed to convert Excel to CSV: {exc}")

        csv_name = f"{Path(filename).stem or 'workbook'}.csv"
        try:
            write_text(build_result_path(req_id, csv_name, base_dir=settings.RESULTS_DIR), csv_text)
        except Exception:  # pragma: no cover - diagnostics helper
            pass

        pipeline_steps.append("excel_to_csv")

        chatgpt_cfg = _get_stage_cfg(pipeline_cfg, "excel", "chatgpt_structured")
        if not _cfg_enabled(chatgpt_cfg, True):
            raise ServiceError(ErrorCode.INTERNAL_ERROR, 503, "Excel ChatGPT disabled via configuration")

        payload = extract_structured_objects(csv_text)
        pipeline_steps.append("chatgpt_structured")
        _persist_json(payload, req_id, "chatgpt_structured.json")
    elif is_doc:
        convert_cfg = _get_stage_cfg(pipeline_cfg, "doc", "doc_to_md")
        if not _cfg_enabled(convert_cfg, True):
            raise ServiceError(ErrorCode.INTERNAL_ERROR, 503, "DOC to Markdown disabled via configuration")

        to_format = "gfm"
        if isinstance(convert_cfg, dict):
            format_value = convert_cfg.get("to_format") or convert_cfg.get("to") or convert_cfg.get("format")
            if isinstance(format_value, str) and format_value.strip():
                to_format = format_value.strip()

            args_value = convert_cfg.get("args")
            extra_args = None
            if isinstance(args_value, (list, tuple)):
                extra_args = [str(arg) for arg in args_value]
            elif isinstance(args_value, str) and args_value.strip():
                extra_args = [arg for arg in args_value.split() if arg]
        else:
            extra_args = None

        try:
            md_text = docx_to_md_text(src_path, to_format=to_format, extra_args=extra_args)
        except Exception as exc:  # noqa: BLE001
            raise ServiceError(ErrorCode.INTERNAL_ERROR, 500, f"Failed to convert DOC to Markdown: {exc}")

        md_name = f"{Path(filename).stem or 'document'}.md"
        try:
            write_text(build_result_path(req_id, md_name, base_dir=settings.RESULTS_DIR), md_text)
        except Exception:  # pragma: no cover - diagnostics helper
            pass

        pipeline_steps.append("doc_to_md")

        chatgpt_cfg = _get_stage_cfg(pipeline_cfg, "doc", "chatgpt_structured")
        if not _cfg_enabled(chatgpt_cfg, True):
            raise ServiceError(ErrorCode.INTERNAL_ERROR, 503, "DOC ChatGPT disabled via configuration")

        payload = extract_structured_objects(md_text)
        pipeline_steps.append("chatgpt_structured")
        _persist_json(payload, req_id, "chatgpt_structured.json")
    elif is_docx:
        convert_cfg = _get_stage_cfg(pipeline_cfg, "docx", "docx_to_md")
        if not _cfg_enabled(convert_cfg, True):
            raise ServiceError(ErrorCode.INTERNAL_ERROR, 503, "DOCX to Markdown disabled via configuration")

        to_format = "gfm"
        if isinstance(convert_cfg, dict):
            format_value = convert_cfg.get("to_format") or convert_cfg.get("to") or convert_cfg.get("format")
            if isinstance(format_value, str) and format_value.strip():
                to_format = format_value.strip()

            args_value = convert_cfg.get("args")
            extra_args = None
            if isinstance(args_value, (list, tuple)):
                extra_args = [str(arg) for arg in args_value]
            elif isinstance(args_value, str) and args_value.strip():
                extra_args = [arg for arg in args_value.split() if arg]
        else:
            extra_args = None

        try:
            md_text = docx_to_md_text(src_path, to_format=to_format, extra_args=extra_args)
        except Exception as exc:  # noqa: BLE001
            raise ServiceError(ErrorCode.INTERNAL_ERROR, 500, f"Failed to convert DOCX to Markdown: {exc}")

        md_name = f"{Path(filename).stem or 'document'}.md"
        try:
            write_text(build_result_path(req_id, md_name, base_dir=settings.RESULTS_DIR), md_text)
        except Exception:  # pragma: no cover - diagnostics helper
            pass

        pipeline_steps.append("docx_to_md")

        chatgpt_cfg = _get_stage_cfg(pipeline_cfg, "docx", "chatgpt_structured")
        if not _cfg_enabled(chatgpt_cfg, True):
            raise ServiceError(ErrorCode.INTERNAL_ERROR, 503, "DOCX ChatGPT disabled via configuration")

        payload = extract_structured_objects(md_text)
        pipeline_steps.append("chatgpt_structured")
        _persist_json(payload, req_id, "chatgpt_structured.json")
    elif is_ppt:
        convert_cfg = _get_stage_cfg(pipeline_cfg, "ppt", "ppt_to_md")
        if not _cfg_enabled(convert_cfg, True):
            raise ServiceError(ErrorCode.INTERNAL_ERROR, 503, "PPT to Markdown disabled via configuration")

        heading_prefix = "# "
        bullet_prefix = "- "
        include_tables = True
        if isinstance(convert_cfg, dict):
            heading_value = convert_cfg.get("heading_prefix")
            if isinstance(heading_value, str):
                heading_prefix = heading_value
            bullet_value = convert_cfg.get("bullet_prefix")
            if isinstance(bullet_value, str):
                bullet_prefix = bullet_value
            include_tables_value = convert_cfg.get("include_tables")
            if isinstance(include_tables_value, bool):
                include_tables = include_tables_value
            elif isinstance(include_tables_value, (int, float)):
                include_tables = bool(include_tables_value)

        try:
            md_text = ppt_to_md_text(
                src_path,
                heading_prefix=heading_prefix,
                bullet_prefix=bullet_prefix,
                include_tables=include_tables,
            )
        except Exception as exc:  # noqa: BLE001
            raise ServiceError(ErrorCode.INTERNAL_ERROR, 500, f"Failed to convert PPT to Markdown: {exc}")

        md_name = f"{Path(filename).stem or 'presentation'}.md"
        try:
            write_text(build_result_path(req_id, md_name, base_dir=settings.RESULTS_DIR), md_text)
        except Exception:  # pragma: no cover - diagnostics helper
            pass

        pipeline_steps.append("ppt_to_md")

        chatgpt_cfg = _get_stage_cfg(pipeline_cfg, "ppt", "chatgpt_structured")
        if not _cfg_enabled(chatgpt_cfg, True):
            raise ServiceError(ErrorCode.INTERNAL_ERROR, 503, "PPT ChatGPT disabled via configuration")

        payload = extract_structured_objects(md_text)
        pipeline_steps.append("chatgpt_structured")
        _persist_json(payload, req_id, "chatgpt_structured.json")
    elif fmt_key == "txt":
        txt_path = Path(src_path)
        try:
            text_content = txt_path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            text_content = txt_path.read_text(encoding="utf-8", errors="ignore")
        except Exception as exc:  # noqa: BLE001
            raise ServiceError(ErrorCode.INTERNAL_ERROR, 500, f"Failed to read TXT file: {exc}") from exc

        pipeline_steps.append("txt_read")

        chatgpt_cfg = _get_stage_cfg(pipeline_cfg, "txt", "chatgpt_structured")
        if not _cfg_enabled(chatgpt_cfg, True):
            raise ServiceError(ErrorCode.INTERNAL_ERROR, 503, "TXT ChatGPT disabled via configuration")

        payload = extract_structured_objects(text_content)
        pipeline_steps.append("chatgpt_structured")
        _persist_json(payload, req_id, "chatgpt_structured.json")
    elif treat_as_pdf:
        if fmt_key == "pdf":
            pdf_path = str(src_path)
        else:
            pdf_stage_cfg = _get_stage_cfg(pipeline_cfg, fmt_key, "pdf_conversion")
            if not _cfg_enabled(pdf_stage_cfg, True):
                raise ServiceError(ErrorCode.PDF_CONVERSION_ERROR, 503, "PDF conversion disabled via configuration")

            try:
                pdf_path = to_pdf(str(src_path), settings.PDF_TMP_DIR, format_cfg, pdf_stage_cfg)
            except ServiceError:
                raise
            except Exception as e:  # noqa: BLE001
                raise ServiceError(ErrorCode.PDF_CONVERSION_ERROR, 422, f"Failed to convert to PDF: {e}")
            pipeline_steps.append("pdf_conversion")

        try:
            pdf_src = Path(pdf_path)
            if pdf_src.exists():
                pdf_dest = build_result_path(req_id, pdf_src.name, base_dir=settings.RESULTS_DIR)
                if str(pdf_dest) != str(pdf_src):
                    write_bytes(pdf_dest, pdf_src.read_bytes())
        except Exception:  # pragma: no cover - diagnostics helper
            pass

        stage_key = "pdf"
        images_stage_cfg = _get_stage_cfg(pipeline_cfg, stage_key, "pdf_to_images")
        if not _cfg_enabled(images_stage_cfg, True):
            raise ServiceError(ErrorCode.INTERNAL_ERROR, 503, "PDF to images disabled via configuration")

        dpi_value = images_stage_cfg.get("dpi")
        try:
            dpi = int(dpi_value) if dpi_value is not None else 150
        except (TypeError, ValueError):
            dpi = 150

        format_value = images_stage_cfg.get("format") or images_stage_cfg.get("image_format")
        if isinstance(format_value, str) and format_value.strip():
            image_format = format_value.strip().lower()
        else:
            image_format = "png"

        poppler_override = images_stage_cfg.get("poppler_path")
        if isinstance(poppler_override, str) and poppler_override.strip():
            poppler_path = poppler_override.strip()
        else:
            poppler_path = settings.POPPLER_PATH

        pages_dir = Path(settings.RESULTS_DIR) / req_id / "pdf_pages"
        try:
            page_images = pdf_to_images(
                pdf_path,
                str(pages_dir),
                dpi=dpi,
                image_format=image_format,
                poppler_path=poppler_path,
            )
        except ServiceError:
            raise
        except Exception as exc:  # noqa: BLE001
            raise ServiceError(ErrorCode.PDF_CONVERSION_ERROR, 422, f"Failed to rasterize PDF: {exc}") from exc

        pipeline_steps.append("pdf_to_images")

        vision_stage_cfg = _get_stage_cfg(pipeline_cfg, stage_key, "vision_per_page")
        if not _cfg_enabled(vision_stage_cfg, True):
            raise ServiceError(ErrorCode.INTERNAL_ERROR, 503, "PDF vision stage disabled via configuration")

        prompt_override = vision_stage_cfg.get("prompt_path") or vision_stage_cfg.get("prompt")
        if isinstance(prompt_override, str) and prompt_override.strip():
            prompt_override = prompt_override.strip()
        else:
            prompt_override = None

        model_override = vision_stage_cfg.get("model")
        if isinstance(model_override, str) and model_override.strip():
            model_override = model_override.strip()
        else:
            model_override = None

        page_payloads: list[dict[str, Any]] = []
        for idx, image_path in enumerate(page_images, start=1):
            page_data = analyze_page_image(image_path, prompt_path=prompt_override, model=model_override)
            if isinstance(page_data, dict) and "page_index" not in page_data:
                page_data["page_index"] = idx
            page_payloads.append(page_data)
            try:
                write_text(
                    build_result_path(req_id, f"pdf_page_{idx:04d}.json", base_dir=settings.RESULTS_DIR),
                    json.dumps(page_data, ensure_ascii=False, indent=2),
                )
            except Exception:  # pragma: no cover - diagnostics helper
                pass

            try:
                write_text(
                    build_result_path(req_id, "pdf_pages_combined.json", base_dir=settings.RESULTS_DIR),
                    json.dumps(page_payloads, ensure_ascii=False, indent=2),
                )
            except Exception:  # pragma: no cover - diagnostics helper
                pass

            pipeline_steps.append("vision_per_page")

        chatgpt_cfg = _get_stage_cfg(pipeline_cfg, stage_key, "chatgpt_structured")
        if _cfg_enabled(chatgpt_cfg, True):
            payload = extract_structured_objects(page_payloads)
            pipeline_steps.append("chatgpt_structured")
            _persist_json(payload, req_id, "chatgpt_structured.json")
        else:
            payload = {"objects": []}
            pipeline_steps.append("chatgpt_skip")
    else:
            pdf_stage_cfg = _get_stage_cfg(pipeline_cfg, fmt_key, "pdf_conversion")
            pdf_enabled_default = fmt_key != "pdf"
            if _cfg_enabled(pdf_stage_cfg, pdf_enabled_default):
                try:
                    pdf_path = to_pdf(str(src_path), settings.PDF_TMP_DIR, format_cfg, pdf_stage_cfg)
                except ServiceError:
                    raise
                except Exception as e:  # noqa: BLE001
                    raise ServiceError(ErrorCode.PDF_CONVERSION_ERROR, 422, f"Failed to convert to PDF: {e}")
                pipeline_steps.append("pdf_conversion")
            else:
                if ext != "pdf":
                    raise ServiceError(
                        ErrorCode.PDF_CONVERSION_ERROR,
                        503,
                        f"PDF conversion disabled for format '{fmt_key}'",
                    )
                pdf_path = str(src_path)

            try:
                pdf_src = Path(pdf_path)
                if pdf_src.exists():
                    pdf_dest = build_result_path(req_id, pdf_src.name, base_dir=settings.RESULTS_DIR)
                    if str(pdf_dest) != str(pdf_src):
                        write_bytes(pdf_dest, pdf_src.read_bytes())
            except Exception:  # pragma: no cover - diagnostics helper
                pass

            if query and query.strip():
                query_text = query
            else:
                qpath = Path(settings.DEFAULT_QUERY_PATH)
                if not qpath.exists():
                    raise ServiceError(ErrorCode.INTERNAL_ERROR, 500, f"Default query not found: {qpath}")
                query_text = qpath.read_text(encoding="utf-8")

            agentql_stage_cfg = _get_stage_cfg(pipeline_cfg, fmt_key, "agentql")
            if _cfg_enabled(agentql_stage_cfg, True):
                agentql_kwargs: dict[str, Any] = {}
                mode_value = agentql_stage_cfg.get("mode")
                if isinstance(mode_value, str) and mode_value.strip():
                    agentql_kwargs["mode"] = mode_value.strip()
                else:
                    agentql_kwargs["mode"] = "standard"
                timeout_value = agentql_stage_cfg.get("timeout_sec")
                if timeout_value is not None:
                    try:
                        agentql_kwargs["timeout_sec"] = float(timeout_value)
                    except (TypeError, ValueError):
                        pass

                agentql_mode_meta = agentql_kwargs.get("mode")

                try:
                    aql_resp = run_agentql(pdf_path, query_text, **agentql_kwargs)
                except ServiceError:
                    raise
                except Exception as e:  # noqa: BLE001
                    raise ServiceError(ErrorCode.AGENTQL_ERROR, 424, f"AgentQL failed: {e}")

                pipeline_steps.append("agentql")

                if isinstance(aql_resp, dict):
                    _persist_json(aql_resp, req_id, "agentql.json")
                    payload = aql_resp
                else:
                    _persist_json({"data": aql_resp}, req_id, "agentql.json")
                    payload = {}
            else:
                payload = {"objects": []}
                pipeline_steps.append("agentql_skip")


    objects, pending_questions = normalize_agentql_payload(payload, rules)

    rows = flatten_objects_to_listings(objects, rules, request_id=req_id, source_file=filename)

    excel_stage_cfg = _get_stage_cfg(pipeline_cfg, "postprocess", "excel_export")
    excel_enabled = _cfg_enabled(excel_stage_cfg, True)

    columns: list[object] = []
    xlsx_bytes: bytes | None = None
    export_path = None
    if excel_enabled:
        raw_columns = rules["output"]["listing_columns"]
        for col in raw_columns:
            if isinstance(col, str) and "|" in col:
                key, header = [part.strip() for part in col.split("|", 1)]
                if not key:
                    continue
                columns.append((key, header or key))
            else:
                columns.append(col)
        export_path = build_result_path(req_id, "listings.xlsx", base_dir=settings.RESULTS_DIR)
        xlsx_bytes = build_xlsx(rows, columns=columns)
        write_bytes(export_path, xlsx_bytes)
        pipeline_steps.append("excel_export")

    elapsed_ms = int((time.perf_counter() - start_ts) * 1000)
    listings_total = len(rows)

    pipeline_marker = "_".join(step for step in pipeline_steps if step)

    excel_url = None
    if excel_enabled and settings.BASE_URL and export_path is not None:
        base = settings.BASE_URL.rstrip("/")
        excel_url = f"{base}/results/{req_id}/listings.xlsx"

    out_mode = (output or "excel").lower()
    if out_mode == "excel":
        if not excel_enabled or xlsx_bytes is None:
            raise ServiceError(ErrorCode.INTERNAL_ERROR, 503, "Excel export disabled via configuration")
        status_payload = [
            {
                "message": "✅ Готово: сводная таблица. Пожалуйста проверьте корректность распознавания и отправьте обратно в бота, чтобы опубликовать в CRM"
            }
        ]
        status_header = base64.b64encode(json.dumps(status_payload, ensure_ascii=False).encode("utf-8")).decode("ascii")
        headers = {
            "Content-Disposition": 'attachment; filename="listings.xlsx"',
            "X-UD-Status": status_header,
        }
        return Response(
            content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    body = {
        "request_id": req_id,
        "items_count": len(rows),
        "objects": None,
        "excel_url": excel_url,
        "pending_questions": pending_questions,
        "meta": {
            "source_file": os.path.basename(filename),
            "listings_total": listings_total,
            "timing_ms": elapsed_ms,
            "pipeline": pipeline_marker,
            "agentql_mode": agentql_mode_meta,
        },
    }

    return body
def _excel_has_listing_headers(path: Path, rules: Mapping[str, Any]) -> bool:
    output_cfg = rules.get("output") if isinstance(rules, Mapping) else None
    header_map: dict[str, str] = {}
    if isinstance(output_cfg, Mapping):
        raw_columns = output_cfg.get("listing_columns")
        if isinstance(raw_columns, (list, tuple)):
            for col in raw_columns:
                key: str | None = None
                header: str | None = None
                if isinstance(col, str):
                    if "|" in col:
                        k, h = col.split("|", 1)
                        key = k.strip()
                        header = h.strip()
                    else:
                        key = col.strip()
                        header = key
                elif isinstance(col, Mapping):
                    raw_key = col.get("key") or col.get("id")
                    if raw_key is not None:
                        key = str(raw_key).strip()
                        raw_header = col.get("title") or col.get("header") or col.get("name")
                        header = str(raw_header).strip() if raw_header is not None else key
                elif isinstance(col, (list, tuple)) and col:
                    key = str(col[0]).strip()
                    header = str(col[1]).strip() if len(col) > 1 and col[1] is not None else key
                if key and header:
                    header_map[key] = header

    required_keys = ["building_name", "area_sqm", "use_type_norm"]
    expected_headers = [header_map.get(k, k).strip().lower() for k in required_keys]

    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            iterator = ws.iter_rows(min_row=1, max_row=1, values_only=True)
            header_row = next(iterator, None)
            if not header_row:
                return False
            actual = {str(cell).strip().lower() for cell in header_row if cell is not None and str(cell).strip()}
            return all(h in actual for h in expected_headers)
        finally:
            wb.close()
    except Exception:
        return False
