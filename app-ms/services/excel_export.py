from __future__ import annotations

from io import BytesIO
from typing import Mapping, Sequence

from openpyxl import Workbook


def _normalize_columns(columns: Sequence[object] | None, rows: list[dict]) -> tuple[list[str], list[str]]:
    """Return parallel lists of data keys and header titles."""
    if columns is None:
        keys = list(rows[0].keys()) if rows else []
        headers = list(keys)
        return keys, headers

    keys: list[str] = []
    headers: list[str] = []

    for col in columns:
        key: str
        header: str

        if isinstance(col, str):
            if "|" in col:
                raw_key, raw_header = col.split("|", 1)
                key = raw_key.strip()
                header = raw_header.strip() or key
            else:
                key = col.strip()
                header = key
        elif isinstance(col, Mapping):
            raw_key = col.get("key") or col.get("id")
            if raw_key is None:
                continue
            key = str(raw_key)
            raw_header = col.get("title") or col.get("header") or col.get("name")
            header = str(raw_header) if raw_header is not None else key
        elif isinstance(col, (tuple, list)) and len(col):
            key = str(col[0])
            header = str(col[1]) if len(col) > 1 and col[1] is not None else key
        else:
            key = str(col)
            header = key

        if not key:
            continue
        keys.append(key)
        headers.append(header)

    return keys, headers


def build_xlsx(rows: list[dict], columns: Sequence[object] | None = None) -> bytes:
    """
    Build an Excel workbook from rows using the exact column order provided.

    - rows: list of dictionaries (values may be None/int/float/str)
    - columns: ordered list of column descriptors; supports:
      * plain key strings ("field_name")
      * pipe-delimited "field|Header" strings
      * (key, header) tuples or lists
      * mappings with the "key" field and optional "title"/"header"/"name" for display
    - Applies:
      - header row freeze (A2)
      - auto filter across the used range
      - simple number formats: int -> "0", float -> "0.00"
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Result"

    rows = list(rows or [])
    keys, headers = _normalize_columns(columns, rows)

    if not rows:
        if headers:
            ws.append(headers)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    if not headers:
        headers = list(rows[0].keys())
        keys = headers

    ws.append(headers)

    for r in rows:
        row_vals = [r.get(k) for k in keys]
        ws.append(row_vals)

    # Apply simple number formats (no currency symbols)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            val = cell.value
            if isinstance(val, int):
                cell.number_format = "0"
            elif isinstance(val, float):
                cell.number_format = "0.00"

    if headers:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
