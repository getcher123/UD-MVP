from __future__ import annotations

from io import StringIO
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


def _format_row(values: Iterable[object]) -> str:
    cells: list[str] = []
    for value in values:
        if value is None:
            cells.append("")
            continue
        if isinstance(value, float):
            text = ("%s" % value).rstrip("0").rstrip(".") if value == int(value) else str(value)
        else:
            text = str(value)
        text = text.replace('"', '""')
        cells.append(f'"{text}"')
    return ",".join(cells)


def excel_to_csv_text(path: str | Path) -> str:
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    buffer = StringIO()
    try:
        first_sheet = True
        for sheet in wb.worksheets:
            if not first_sheet:
                buffer.write("\n")
            first_sheet = False
            buffer.write(f"# sheet: {sheet.title}\n")
            for row in sheet.iter_rows(values_only=True):
                buffer.write(_format_row(row))
                buffer.write("\n")
    finally:
        wb.close()
    return buffer.getvalue().strip()


__all__ = ["excel_to_csv_text"]
