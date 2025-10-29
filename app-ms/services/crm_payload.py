from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from typing import Any, Dict, Iterable, Mapping, Optional

from openpyxl import load_workbook

from core.errors import ErrorCode, ServiceError

_NUMERIC_FIELDS = {
    "area_sqm",
    "divisible_from_sqm",
    "rent_rate_year_sqm_base",
    "opex_year_per_sqm",
    "rent_month_total_gross",
    "sale_price_per_sqm",
}
_UNCERTAIN_SPLITTER = re.compile(r"[;\n]+")


def prepare_crm_payload(
    excel_path: str,
    request_id: str,
    source_file: str,
    rules: Mapping[str, Any],
) -> Dict[str, Any]:
    listings = _read_listings(excel_path, request_id, source_file, rules)
    if not listings:
        raise ServiceError(ErrorCode.CRM_SYNC_ERROR, 400, "listings.xlsx does not contain any rows")

    received_at = datetime.now(tz=timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    payload: Dict[str, Any] = {
        "request_id": request_id,
        "source_file": source_file,
        "received_at": received_at,
        "listings": listings,
        "meta": {
            "listings_total": len(listings),
            "origin": "app-ms",
        },
    }
    return payload


def _read_listings(
    excel_path: str,
    request_id: str,
    source_file: str,
    rules: Mapping[str, Any],
) -> list[dict[str, Any]]:
    header_map, key_map = _build_header_lookup(rules)

    wb = load_workbook(excel_path, data_only=True, read_only=True)
    try:
        sheet = wb.active
        iterator = sheet.iter_rows(min_row=1, values_only=True)
        try:
            headers = next(iterator)
        except StopIteration:
            return []

        resolved_keys = [_resolve_key(h, header_map, key_map) for h in headers]
        listings: list[dict[str, Any]] = []

        for row_index, row_values in enumerate(iterator, start=2):
            if _is_row_empty(row_values):
                continue

            listing: dict[str, Any] = {}
            for idx, key in enumerate(resolved_keys):
                if not key or idx >= len(row_values):
                    continue
                value = _normalize_value(key, row_values[idx])
                if value is None:
                    continue
                listing[key] = value

            listing.setdefault("source_file", source_file)
            listing.setdefault("request_id", request_id)
            listing["uncertain_parameters"] = _ensure_uncertain_list(listing.get("uncertain_parameters"))

            if "building_name" not in listing or not str(listing["building_name"]).strip():
                raise ServiceError(ErrorCode.CRM_SYNC_ERROR, 400, f"Row {row_index} is missing building_name")

            if "area_sqm" not in listing or listing["area_sqm"] is None:
                raise ServiceError(ErrorCode.CRM_SYNC_ERROR, 400, f"Row {row_index} is missing area_sqm")

            if not isinstance(listing["building_name"], str):
                listing["building_name"] = str(listing["building_name"])

            listings.append(listing)

        return listings
    finally:
        wb.close()


def _build_header_lookup(rules: Mapping[str, Any]) -> tuple[dict[str, str], dict[str, str]]:
    output_cfg = rules.get("output") if isinstance(rules, Mapping) else None
    columns: Iterable[Any] = ()
    if isinstance(output_cfg, Mapping):
        raw_columns = output_cfg.get("listing_columns")
        if isinstance(raw_columns, Iterable):
            columns = raw_columns

    header_map: dict[str, str] = {}
    key_map: dict[str, str] = {}

    for col in columns:
        key: Optional[str] = None
        header: Optional[str] = None

        if isinstance(col, str):
            if "|" in col:
                raw_key, raw_header = col.split("|", 1)
                key = raw_key.strip()
                header = raw_header.strip()
            else:
                key = col.strip()
                header = key
        elif isinstance(col, Mapping):
            raw_key = col.get("key") or col.get("id")
            if raw_key is None:
                continue
            key = str(raw_key).strip()
            raw_header = col.get("title") or col.get("header") or col.get("name")
            header = str(raw_header).strip() if raw_header is not None else key
        elif isinstance(col, (list, tuple)) and col:
            key = str(col[0]).strip()
            if len(col) > 1 and col[1] is not None:
                header = str(col[1]).strip()
            else:
                header = key
        else:
            key = str(col).strip()
            header = key

        if not key:
            continue

        key_lower = key.lower()
        key_map[key_lower] = key

        if header:
            header_map[header.lower()] = key

    return header_map, key_map


def _resolve_key(header: Any, header_map: Mapping[str, str], key_map: Mapping[str, str]) -> Optional[str]:
    if header is None:
        return None
    header_str = str(header).strip()
    if not header_str:
        return None

    key = header_map.get(header_str.lower())
    if key:
        return key

    return key_map.get(header_str.lower())


def _is_row_empty(values: Iterable[Any]) -> bool:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return False
    return True


def _normalize_value(key: str, value: Any) -> Any:
    if value is None:
        return None

    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        if key in _NUMERIC_FIELDS:
            return _parse_numeric_string(key, text)
        if key == "uncertain_parameters":
            return _ensure_uncertain_list(text)
        return text

    if isinstance(value, (int, float)):
        if key in _NUMERIC_FIELDS:
            return float(value)
        return value

    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:  # pragma: no cover - defensive
            return str(value)

    return value


def _parse_numeric_string(key: str, text: str) -> Optional[float]:
    cleaned = text.replace(" ", "").replace("\u00a0", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        # treat unparseable optional values as missing; mandatory fields validated later
        return None


def _ensure_uncertain_list(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return [item for item in value if not isinstance(item, str) or item.strip()]
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return []
        try:
            parsed = json.loads(text)
        except ValueError:
            parsed_list = [part.strip() for part in _UNCERTAIN_SPLITTER.split(text) if part.strip()]
            return parsed_list
        else:
            if isinstance(parsed, list):
                return [item for item in parsed if not isinstance(item, str) or item.strip()]
            if isinstance(parsed, str) and parsed.strip():
                return [parsed.strip()]
            return []
    return [value]


__all__ = ["prepare_crm_payload"]
