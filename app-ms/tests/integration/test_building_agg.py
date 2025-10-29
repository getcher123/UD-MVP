from __future__ import annotations

from io import BytesIO
from pathlib import Path
import sys

from typing import Any, Dict, List

from openpyxl import load_workbook

# Resolve project root so imports like 'from services.aggregate_buildings import ...' work
root = Path(__file__).resolve()
while root.name not in {"app-ms", "app_ms"} and root.parent != root:
    root = root.parent
sys.path.insert(0, str(root))

from services.aggregate_buildings import group_to_buildings  # type: ignore  # noqa: E402
from services.excel_export import build_xlsx  # type: ignore  # noqa: E402


def _rules() -> Dict[str, Any]:
    return {
        "aggregation": {
            "building": {
                "name": {"compose": "{object_name}{suffix}"},
                "source_files": {"unique_join": " | "},
            }
        },
        "normalization": {
            "use_type": {"canon": ["офис"], "synonyms": {"офис": ["office"]}},
            "fitout_condition": {"canon": ["с отделкой", "под отделку"], "synonyms": {}},
            "vat": {"treat_not_applied": ["усн", "упрощенка", "ндс 5%"]},
            "floor": {
                "drop_tokens": ["этаж", "эт", "э."],
                "map_special": {"basement": ["подвал", "-1"], "socle": ["цоколь"], "mezzanine": ["мезонин"]},
                "multi": {
                    "enabled": True,
                    "split_separators": [",", ";", "/", " и ", "&"],
                    "range_separators": ["-", "–"],
                    "render": {"join_token": "; ", "range_dash": "–", "sort_numeric_first": True, "uniq": True},
                },
            },
        },
        "derivation": {
            "rent_rate_year_sqm_base": {
                "priority": ["direct", "reconstruct_from_month"],
                "reconstruct_from_month": {"respect_vat": True, "respect_opex": True, "vat_fallback": 0.2, "round_decimals": 2},
            },
            "gross_month_total": {"round_decimals": 2},
        },
        "output": {
            "building_columns": [
                "building_id",
                "building_name",
                "object_id",
                "object_name",
                "use_type_set_norm",
                "fitout_condition_mode",
                "delivery_date_earliest",
                "floors_covered_norm",
                "area_sqm_total",
                "listing_count",
                "rent_rate_year_sqm_base_min",
                "rent_rate_year_sqm_base_avg",
                "rent_rate_year_sqm_base_max",
                "rent_vat_norm_mode",
                "opex_year_per_sqm_avg",
                "rent_month_total_gross_avg",
                "sale_price_per_sqm_min",
                "sale_price_per_sqm_avg",
                "sale_price_per_sqm_max",
                "sale_vat_norm_mode",
                "source_files",
                "request_id",
                "uncertain_parameters",
            ]
        },
    }


def test_building_aggregation_and_excel_columns(tmp_path: Path):
    rules = _rules()
    objects: List[Dict[str, Any]] = [
        {
            "object_name": "Объект",
            "buildings": [
                {
                    "building_name": "стр. 1",
                    "listings": [
                        {"floor": "1", "area_sqm": 100, "use_type": "офис"},
                        {"floor": "1-2", "area_sqm": 50, "use_type": "office"},
                    ],
                },
                {
                    "building_name": "корпус 2",
                    "listings": [{"floor": "5", "area_sqm": 10, "use_type": "офис"}],
                },
            ],
        }
    ]

    rows = group_to_buildings(objects, rules, request_id="rid-xyz", source_file="/data/source.pdf")
    assert isinstance(rows, list) and rows
    # two buildings aggregated
    assert len(rows) == 2

    # find building with token 'стр. 1'
    row1 = next(r for r in rows if r["building_id"].startswith("obekt__str-1"))
    assert row1["floors_covered_norm"] == "1–2"
    assert row1["listing_count"] == 2
    assert abs(float(row1["area_sqm_total"]) - 150.0) < 1e-6
    assert row1["building_name"] == "Объект, стр. 1"
    assert row1["building_id"] == "obekt__str-1"

    # Excel columns in exact order
    columns = rules["output"]["building_columns"]
    xlsx = build_xlsx(rows, columns=columns)
    wb = load_workbook(BytesIO(xlsx))
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header == columns

